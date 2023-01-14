#为框架加载数据
from pathlib import Path
import csv
from openpyxl import load_workbook
#---------------当前目录---上一级目录
data_dir = Path(__file__).parent.parent / "data"  #数据存放目录

def read_csv(filename):

    path = data_dir / filename  #数据目录+数据名字，拼接出完整的路径

    with open(path,encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i in reader:
            yield list(i.values())  #yiled 起到返回调用者的效果

def read_excel(filename):
    path = data_dir / filename
    wb = load_workbook(path)   #打开文件
    ws = wb.active   #加载数据，切换sheet页

    yield from ws.iter_rows(min_row=2,values_only=True)  


if __name__ == '__main__':
    #read_csv('ddt_test_login.csv')
    for d in read_excel("ddt_test_new_address.xlsx"):
        print(d)